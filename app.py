"""
域名批量查询系统 v2.2
功能：批量查询、暂停/继续、代理支持、历史记录
"""

import os
import re
import time
import json
import random
import logging
import threading
import sqlite3
import shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file
import whois
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = os.urandom(24)

# 数据目录
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
DB_PATH = os.path.join(DATA_DIR, 'domain_checker.db')

# 确保数据目录存在
os.makedirs(DATA_DIR, exist_ok=True)

# 配置（可动态修改）
config_lock = threading.Lock()
CONFIG = {
    'max_domains_per_batch': 500,
    'rate_limit_delay': 1.0,
    'max_retries': 3,
    'retry_delay': 2,
    'timeout': 15,
    'max_workers': 5,
    'proxy_enabled': False,
    'proxy_url': 'http://127.0.0.1:7897',
    'proxy_auth': None,
    'platform': 'whois',  # 查询平台: whois, whoisxml, rdap
}

# 平台信息
PLATFORMS = {
    'whois': {'name': 'WHOIS标准查询', 'icon': '🔍', 'desc': '使用标准WHOIS协议'},
    'whoisxml': {'name': 'WHOIS XML', 'icon': '🌐', 'desc': '使用WHOIS XML API'},
    'rdap': {'name': 'RDAP安全查询', 'icon': '🛡️', 'desc': '使用RDAP协议，更安全'}
}

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# 全局变量
task_storage = {}
task_lock = threading.Lock()
task_pause_flags = {}

# ============== 数据库初始化 ==============

def init_db():
    """初始化SQLite数据库"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 查询历史记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS query_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id TEXT UNIQUE NOT NULL,
            domains TEXT NOT NULL,
            domain_count INTEGER NOT NULL,
            results_count INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            failed_count INTEGER DEFAULT 0,
            status TEXT DEFAULT 'processing',
            created_at TEXT NOT NULL,
            completed_at TEXT,
            config TEXT
        )
    ''')
    
    # 查询详情表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS query_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id TEXT NOT NULL,
            domain TEXT NOT NULL,
            status TEXT,
            registrar TEXT,
            registration_date TEXT,
            expiration_date TEXT,
            updated_date TEXT,
            name_servers TEXT,
            dnssec TEXT,
            resolved INTEGER,
            block_reason TEXT,
            dns_records TEXT,
            error TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY (task_id) REFERENCES query_history(task_id)
        )
    ''')
    
    conn.commit()
    conn.close()
    logger.info(f"数据库初始化完成: {DB_PATH}")

def save_history(task_id: str, domains: list, status: str = 'processing', completed_at: str = None):
    """保存查询历史"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    success_count = 0
    failed_count = 0
    
    if status == 'completed':
        cursor.execute('SELECT status, COUNT(*) FROM query_results WHERE task_id = ? GROUP BY status', (task_id,))
        for row in cursor.fetchall():
            if row[0] == 'success':
                success_count = row[1]
            else:
                failed_count += row[1]
    
    config_json = json.dumps({k: v for k, v in CONFIG.items() if k != 'proxy_auth'})
    
    cursor.execute('''
        INSERT OR REPLACE INTO query_history 
        (task_id, domains, domain_count, results_count, success_count, failed_count, status, created_at, completed_at, config)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        task_id,
        '\n'.join(domains),
        len(domains),
        success_count + failed_count,
        success_count,
        failed_count,
        status,
        datetime.now().isoformat(),
        completed_at,
        config_json
    ))
    
    conn.commit()
    conn.close()

def save_results(task_id: str, results: list):
    """保存查询结果"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 先删除旧结果
    cursor.execute('DELETE FROM query_results WHERE task_id = ?', (task_id,))
    
    for r in results:
        dns_records = r.get('dns_records', [])
        if isinstance(dns_records, list):
            dns_records = ','.join(dns_records)
        
        cursor.execute('''
            INSERT INTO query_results
            (task_id, domain, status, registrar, registration_date, expiration_date, updated_date, 
             name_servers, dnssec, resolved, block_reason, dns_records, error, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            task_id, r.get('domain', ''), r.get('status', ''), r.get('registrar', ''),
            r.get('registration_date', ''), r.get('expiration_date', ''), r.get('updated_date', ''),
            r.get('name_servers', ''), r.get('dnssec', ''), 
            1 if r.get('resolved') == True else (0 if r.get('resolved') == False else None),
            r.get('block_reason', ''), dns_records, r.get('error', ''),
            datetime.now().isoformat()
        ))
    
    conn.commit()
    conn.close()

def update_history_counts(task_id: str, success_count: int, failed_count: int):
    """更新历史记录的统计"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE query_history 
        SET results_count = ?, success_count = ?, failed_count = ?, status = 'completed', completed_at = ?
        WHERE task_id = ?
    ''', (success_count + failed_count, success_count, failed_count, datetime.now().isoformat(), task_id))
    
    conn.commit()
    conn.close()

def get_history(limit: int = 50):
    """获取查询历史"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT * FROM query_history 
        ORDER BY created_at DESC 
        LIMIT ?
    ''', (limit,))
    
    rows = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return rows

def get_history_detail(task_id: str):
    """获取历史详情"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 获取历史记录
    cursor.execute('SELECT * FROM query_history WHERE task_id = ?', (task_id,))
    row = cursor.fetchone()
    history = dict(row) if row else None
    
    # 获取结果
    cursor.execute('SELECT * FROM query_results WHERE task_id = ? ORDER BY id', (task_id,))
    results = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return history, results

def delete_history(task_id: str):
    """删除历史记录"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM query_results WHERE task_id = ?', (task_id,))
    cursor.execute('DELETE FROM query_history WHERE task_id = ?', (task_id,))
    conn.commit()
    conn.close()

def clear_old_history(days: int = 30):
    """清理旧历史"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cutoff = datetime.now()
    cursor.execute('''
        DELETE FROM query_history 
        WHERE created_at < datetime(?, '-' || ? || ' days')
    ''', (cutoff.isoformat(), days))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    logger.info(f"已清理 {deleted} 条{days}天前的历史记录")
    return deleted

# 初始化数据库
init_db()

# ============== 域名解析 ==============

def parse_domain_input(text: str) -> list:
    """解析输入文本，支持纯域名和URL格式
    
    支持格式:
    - example.com
    - www.example.com
    - https://example.com
    - https://www.example.com
    - http://example.com/path (只取域名部分)
    """
    domains = []
    lines = text.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 提取域名
        domain = extract_domain(line)
        if domain:
            domains.append(domain)
    
    # 去重保持顺序
    seen = set()
    unique_domains = []
    for d in domains:
        if d not in seen:
            seen.add(d)
            unique_domains.append(d)
    
    return unique_domains

def extract_domain(text: str) -> str:
    """从URL或纯域名中提取域名"""
    text = text.strip().lower()
    
    # 已经是纯域名
    if is_valid_domain(text):
        return text
    
    # 尝试从URL中提取
    # 匹配协议 + 可能的www + 域名 + 可选路径
    patterns = [
        r'https?://(?:www\.)?([^/]+)',  # https://www.example.com/path
        r'//(?:www\.)?([^/]+)',           # //www.example.com/path
        r'(?:www\.)?([a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,})',  # 兜底
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            domain = match.group(1).strip()
            if is_valid_domain(domain):
                return domain
    
    return None

def is_valid_domain(domain: str) -> bool:
    """验证域名格式"""
    pattern = r'^([a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}$'
    return bool(re.match(pattern, domain.strip()))

# ============== WHOIS查询 ==============

def get_proxies() -> dict:
    """获取代理配置"""
    with config_lock:
        if not CONFIG.get('proxy_enabled') or not CONFIG.get('proxy_url'):
            return None
    
    proxies = {
        'http': CONFIG['proxy_url'],
        'https': CONFIG['proxy_url']
    }
    return proxies

def check_domain_resolved(domain: str) -> dict:
    """检查域名是否被停止解析"""
    result = {'resolved': None, 'dns_records': [], 'http_accessible': None, 'block_reason': None}
    
    try:
        import dns.resolver
        resolver = dns.resolver.Resolver()
        resolver.timeout = 5
        resolver.lifetime = 5
        
        try:
            answers = resolver.resolve(domain, 'A')
            result['resolved'] = True
            result['dns_records'] = [str(rdata) for rdata in answers]
        except dns.resolver.NXDOMAIN:
            result['resolved'] = False
            result['block_reason'] = '域名不存在'
        except dns.resolver.NoAnswer:
            result['resolved'] = False
            result['block_reason'] = '无DNS记录'
        except dns.resolver.NoNameservers:
            result['resolved'] = False
            result['block_reason'] = '无权威DNS服务器'
        except Exception as e:
            result['resolved'] = None
            result['block_reason'] = str(e)[:50]
        
        if result['resolved']:
            try:
                import urllib.request
                url = f'http://{domain}'
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                urllib.request.urlopen(req, timeout=5)
                result['http_accessible'] = True
            except:
                result['http_accessible'] = False
    except:
        result['block_reason'] = '缺少依赖'
    
    return result

def query_whois_with_retry(domain: str) -> dict:
    """带重试机制的WHOIS查询"""
    with config_lock:
        cfg = CONFIG.copy()
    
    last_error = None
    
    for attempt in range(cfg['max_retries']):
        # 检查暂停
        task_id = getattr(threading.current_thread(), 'task_id', None)
        if task_id and task_pause_flags.get(task_id, False):
            while task_pause_flags.get(task_id, False):
                time.sleep(0.5)
        
        try:
            time.sleep(cfg['rate_limit_delay'] + random.uniform(0.1, 0.3))
            
            w = whois.whois(domain)
            
            result = {
                'domain': domain, 'status': 'success',
                'registrar': None, 'registration_date': None, 'expiration_date': None,
                'updated_date': None, 'name_servers': None, 'dnssec': None, 'error': None
            }
            
            if w:
                if w.registrar:
                    result['registrar'] = str(w.registrar)[:100]
                
                for field, key in [('creation_date', 'registration_date'), 
                                   ('expiration_date', 'expiration_date'),
                                   ('updated_date', 'updated_date')]:
                    val = getattr(w, field, None)
                    if val:
                        if isinstance(val, list): val = val[0]
                        if hasattr(val, 'strftime'):
                            result[key] = val.strftime('%Y-%m-%d')
                
                if w.name_servers:
                    ns = w.name_servers if isinstance(w.name_servers, list) else [w.name_servers]
                    result['name_servers'] = ', '.join([str(n)[:30] for n in ns[:5]])
                
                if w.dnssec:
                    result['dnssec'] = str(w.dnssec)[:50]
            
            logger.info(f"[{domain}] WHOIS成功")
            return result
            
        except whois.exceptions.PywhoisError as e:
            last_error = f'WHOIS解析错误: {str(e)[:80]}'
            logger.warning(f"[{domain}] {last_error}")
        except whois.exceptions.WhoisDomainNotFoundError as e:
            last_error = f'域名不存在: {str(e)[:50]}'
            logger.warning(f"[{domain}] {last_error}")
        except whois.exceptions.WhoisQuotaExceededError as e:
            last_error = f'查询配额超限: {str(e)[:50]}'
            logger.warning(f"[{domain}] {last_error}")
        except Exception as e:
            last_error = f'查询异常: {str(e)[:80]}'
            logger.warning(f"[{domain}] {last_error}")
        
        if attempt < cfg['max_retries'] - 1:
            wait_time = cfg['retry_delay'] * (attempt + 1) + random.uniform(0, 1)
            logger.info(f"[{domain}] 重试 ({attempt + 2}/{cfg['max_retries']}), 等待 {wait_time:.1f}s")
            time.sleep(wait_time)
    
    logger.error(f"[{domain}] 查询失败: {last_error}")
    return {
        'domain': domain, 'status': 'failed',
        'registrar': None, 'registration_date': None, 'expiration_date': None,
        'updated_date': None, 'name_servers': None, 'dnssec': None,
        'error': last_error
    }

def process_single_domain(domain: str, task_id: str = None) -> dict:
    """处理单个域名"""
    domain = domain.strip().lower()
    logger.info(f"[{domain}] 开始处理")
    
    if not is_valid_domain(domain):
        logger.warning(f"[{domain}] 格式无效")
        return {
            'domain': domain, 'status': 'invalid',
            'registrar': None, 'registration_date': None, 'expiration_date': None,
            'updated_date': None, 'name_servers': None, 'dnssec': None,
            'error': '域名格式无效', 'resolved': None, 'block_reason': None, 'dns_records': None
        }
    
    whois_result = query_whois_with_retry(domain)
    
    resolve_result = {'resolved': None, 'block_reason': None, 'dns_records': None}
    if whois_result['status'] == 'success':
        resolve_result = check_domain_resolved(domain)
        if resolve_result['resolved']:
            logger.info(f"[{domain}] DNS正常: {resolve_result['dns_records'][:1]}")
        else:
            logger.warning(f"[{domain}] DNS异常: {resolve_result['block_reason']}")
    
    result = {**whois_result, **resolve_result}
    
    # 更新任务
    if task_id:
        with task_lock:
            if task_id in task_storage:
                task_storage[task_id]['refresh'] = True
                
                found = False
                for i, r in enumerate(task_storage[task_id]['results']):
                    if r['domain'] == domain:
                        task_storage[task_id]['results'][i] = result
                        found = True
                        break
                
                if not found:
                    task_storage[task_id]['results'].append(result)
                    task_storage[task_id]['completed'] += 1
                
                # 日志
                log_level = 'info'
                if result['status'] != 'success': log_level = 'error'
                elif resolve_result['resolved'] == False: log_level = 'warn'
                
                task_storage[task_id]['logs'].append({
                    'time': datetime.now().strftime('%H:%M:%S'),
                    'level': log_level,
                    'message': f"{'✓' if result['status'] == 'success' else '✗'} {domain}" + 
                              (f" [已停止解析]" if resolve_result['resolved'] == False else "") +
                              (f" - {result['error']}" if result['error'] else "")
                })
    
    return result

def process_domains_async(domains: list, task_id: str):
    """异步处理域名"""
    total = len(domains)
    task_pause_flags[task_id] = False
    
    logger.info(f"[任务{task_id}] 开始处理 {total} 个域名")
    
    with task_lock:
        task_storage[task_id]['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'level': 'info',
            'message': f'任务开始，共 {total} 个域名'
        })
    
    # 保存初始历史
    save_history(task_id, domains, 'processing')
    
    # 直接使用线程处理
    threads = []
    for d in domains:
        while task_pause_flags.get(task_id, False):
            time.sleep(0.3)
        
        t = threading.Thread(target=process_single_domain, args=(d, task_id))
        t.daemon = True
        t.start()
        threads.append(t)
        time.sleep(0.1)
    
    # 等待所有线程完成
    for t in threads:
        t.join()
    
    # 保存结果到数据库
    with task_lock:
        results = list(task_storage[task_id]['results'])
        success_count = sum(1 for r in results if r['status'] == 'success')
        failed_count = len(results) - success_count
        
        task_pause_flags.pop(task_id, None)
        task_storage[task_id]['status'] = 'completed'
        task_storage[task_id]['completed_at'] = datetime.now().isoformat()
        task_storage[task_id]['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'level': 'info',
            'message': f'任务完成！共 {len(results)} 条结果'
        })
    
    # 在锁外保存到数据库
    save_results(task_id, results)
    update_history_counts(task_id, success_count, failed_count)
    
    logger.info(f"[任务{task_id}] 全部完成")

def generate_task_id():
    return datetime.now().strftime('%m%d%H%M%S') + str(random.randint(100, 999))

# ============== 导出 ==============

def create_export_file(results: list, format: str = 'csv', filter_type: str = 'all') -> tuple:
    filtered = results
    if filter_type == 'success':
        filtered = [r for r in results if r['status'] == 'success']
    elif filter_type == 'normal':
        filtered = [r for r in results if r['status'] == 'success' and r.get('resolved') != False]
    elif filter_type == 'failed':
        filtered = [r for r in results if r['status'] != 'success']
    elif filter_type == 'blocked':
        filtered = [r for r in results if r.get('resolved') == False]
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if format == 'csv':
        return create_csv(filtered, timestamp), f'domain_report_{timestamp}.csv'
    else:
        return create_xlsx(filtered, timestamp), f'domain_report_{timestamp}.xlsx'

def create_csv(results: list, timestamp: str) -> str:
    import csv
    filepath = os.path.join('/tmp', f'domain_report_{timestamp}.csv')
    headers = ['域名', '状态', '注册商', '注册日期', '过期日期', '更新时间', 
               'DNS服务器', 'DNSSEC', '解析状态', 'DNS记录', '封禁原因', '错误备注']
    
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for r in results:
            status_map = {'success': '查询成功', 'failed': '查询失败', 'invalid': '格式无效'}
            resolved_map = {True: '正常解析', False: '已停止解析', None: '未知'}
            dns_records = r.get('dns_records', [])
            if isinstance(dns_records, list): dns_records = ', '.join(dns_records)
            
            writer.writerow([
                r.get('domain', ''), status_map.get(r.get('status', ''), ''),
                r.get('registrar', ''), r.get('registration_date', ''),
                r.get('expiration_date', ''), r.get('updated_date', ''),
                r.get('name_servers', ''), r.get('dnssec', ''),
                resolved_map.get(r.get('resolved'), ''), dns_records,
                r.get('block_reason', ''), r.get('error', '')
            ])
    
    return filepath

def create_xlsx(results: list, timestamp: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "域名查询结果"
    
    headers = ['域名', '状态', '注册商', '注册日期', '过期日期', '更新时间', 
               'DNS服务器', 'DNSSEC', '解析状态', 'DNS记录', '封禁原因', '错误备注']
    
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, r in enumerate(results, 2):
        status_map = {'success': '查询成功', 'failed': '查询失败', 'invalid': '格式无效'}
        resolved_map = {True: '正常解析', False: '已停止解析', None: '未知'}
        dns_records = r.get('dns_records', [])
        if isinstance(dns_records, list): dns_records = ', '.join(dns_records)
        
        for col, val in enumerate([
            r.get('domain', ''), status_map.get(r.get('status', ''), ''),
            r.get('registrar', ''), r.get('registration_date', ''),
            r.get('expiration_date', ''), r.get('updated_date', ''),
            r.get('name_servers', ''), r.get('dnssec', ''),
            resolved_map.get(r.get('resolved'), ''), dns_records,
            r.get('block_reason', ''), r.get('error', '')
        ], 1):
            ws.cell(row=row_idx, column=col, value=val)
    
    for i, w in enumerate([30, 12, 20, 12, 12, 12, 35, 15, 12, 25, 25, 35], 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w
    
    filepath = os.path.join('/tmp', f'domain_report_{timestamp}.xlsx')
    wb.save(filepath)
    return filepath

# ============== Flask路由 ==============

@app.route('/')
def index():
    return render_template('index.html', config=CONFIG)

@app.route('/api/config', methods=['GET', 'POST'])
def api_config():
    global CONFIG
    if request.method == 'POST':
        data = request.get_json()
        with config_lock:
            for key in CONFIG:
                if key in data:
                    val = data[key]
                    if key in ['max_domains_per_batch', 'max_retries', 'timeout', 'max_workers']:
                        val = int(val)
                    elif key in ['rate_limit_delay', 'retry_delay']:
                        val = float(val)
                    elif key in ['proxy_enabled']:
                        val = bool(val)
                    elif key == 'platform':
                        if val in PLATFORMS:
                            val = val
                        else:
                            val = 'whois'
                    CONFIG[key] = val
        logger.info(f"配置已更新: platform={CONFIG['platform']}, proxy_enabled={CONFIG['proxy_enabled']}")
        return jsonify({'message': '配置已更新', 'config': CONFIG, 'platforms': PLATFORMS})
    return jsonify({**CONFIG, 'platforms': PLATFORMS})

@app.route('/api/query', methods=['POST'])
def api_query():
    data = request.get_json()
    input_text = data.get('domains', '')
    platform = data.get('platform', 'whois')
    
    # 解析输入（支持URL格式）
    domains = parse_domain_input(input_text)
    
    if not domains:
        return jsonify({'error': '请输入有效域名'}), 400
    
    with config_lock:
        max_batch = CONFIG['max_domains_per_batch']
        # 如果传入了平台参数，先更新配置
        if platform in PLATFORMS:
            CONFIG['platform'] = platform
    
    if len(domains) > max_batch:
        return jsonify({'error': f'单次查询最多{max_batch}个域名'}), 400
    
    task_id = generate_task_id()
    
    with task_lock:
        task_storage[task_id] = {
            'status': 'processing', 'total': len(domains), 'completed': 0,
            'results': [], 'logs': [], 'refresh': False,
            'created_at': datetime.now().isoformat(), 'completed_at': None,
            'platform': platform
        }
    
    logger.info(f"[任务{task_id}] 创建，使用平台: {PLATFORMS.get(platform, {}).get('name', platform)}")
    
    thread = threading.Thread(target=process_domains_async, args=(domains, task_id))
    thread.daemon = True
    thread.start()
    
    return jsonify({
        'task_id': task_id, 'total': len(domains), 'platform': platform,
        'message': f'任务已创建，使用{PLATFORMS.get(platform, {}).get("name", platform)}处理 {len(domains)} 个域名'
    })

@app.route('/api/status/<task_id>')
def api_status(task_id):
    with task_lock:
        task = task_storage.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    return jsonify({
        'status': task['status'], 'total': task['total'],
        'completed': task['completed'],
        'progress': round(task['completed'] / task['total'] * 100, 1) if task['total'] > 0 else 0,
        'paused': task_pause_flags.get(task_id, False)
    })

@app.route('/api/results/<task_id>')
def api_results(task_id):
    with task_lock:
        task = task_storage.get(task_id)
    
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    
    log_level = request.args.get('log_level', 'all')
    logs = task['logs']
    if log_level != 'all':
        logs = [l for l in logs if l.get('level') == log_level]
    
    refresh = task.get('refresh', False)
    if refresh:
        with task_lock:
            task_storage[task_id]['refresh'] = False
    
    return jsonify({
        'status': task['status'], 'results': task['results'],
        'total': task['total'], 'completed': task['completed'],
        'logs': logs[-100:], 'refresh': refresh,
        'paused': task_pause_flags.get(task_id, False)
    })

@app.route('/api/pause/<task_id>', methods=['POST'])
def api_pause(task_id):
    task_pause_flags[task_id] = True
    with task_lock:
        if task_id in task_storage:
            task_storage[task_id]['logs'].append({
                'time': datetime.now().strftime('%H:%M:%S'),
                'level': 'warn',
                'message': '⏸ 任务已暂停'
            })
    return jsonify({'message': '已暂停', 'paused': True})

@app.route('/api/resume/<task_id>', methods=['POST'])
def api_resume(task_id):
    task_pause_flags[task_id] = False
    with task_lock:
        if task_id in task_storage:
            task_storage[task_id]['logs'].append({
                'time': datetime.now().strftime('%H:%M:%S'),
                'level': 'info',
                'message': '▶ 任务已继续'
            })
    return jsonify({'message': '已继续', 'paused': False})

@app.route('/api/retry/<task_id>', methods=['POST'])
def api_retry(task_id):
    data = request.get_json()
    domains = data.get('domains', [])
    if not domains:
        return jsonify({'error': '请选择域名'}), 400
    
    with task_lock:
        task = task_storage.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    
    with task_lock:
        task['refresh'] = True
        task['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'level': 'info',
            'message': f'🔄 开始重新查询 {len(domains)} 个域名'
        })
    
    def do_retry():
        for domain in domains:
            process_single_domain(domain, task_id)
    
    thread = threading.Thread(target=do_retry)
    thread.daemon = True
    thread.start()
    
    return jsonify({'message': f'正在重新查询 {len(domains)} 个域名'})

@app.route('/api/retry-failed/<task_id>', methods=['POST'])
def api_retry_failed(task_id):
    with task_lock:
        task = task_storage.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    
    failed_domains = [r['domain'] for r in task['results'] if r['status'] != 'success']
    if not failed_domains:
        return jsonify({'message': '没有需要重试的域名'})
    
    with task_lock:
        task['refresh'] = True
        task['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'level': 'info',
            'message': f'🔄 开始重试 {len(failed_domains)} 个失败域名'
        })
    
    def do_retry():
        for domain in failed_domains:
            process_single_domain(domain, task_id)
    
    thread = threading.Thread(target=do_retry)
    thread.daemon = True
    thread.start()
    
    return jsonify({'message': f'正在重试 {len(failed_domains)} 个域名'})

@app.route('/api/export/<task_id>')
def api_export(task_id):
    format_type = request.args.get('format', 'csv')
    filter_type = request.args.get('filter', 'all')
    selected = request.args.get('selected', '')
    
    # 优先从内存获取，否则从数据库
    with task_lock:
        task = task_storage.get(task_id)
    
    if task:
        results = task['results']
    else:
        # 从数据库获取
        _, results = get_history_detail(task_id)
        results = [{'domain': r['domain'], 'status': r['status'], 'registrar': r['registrar'],
                   'registration_date': r['registration_date'], 'expiration_date': r['expiration_date'],
                   'updated_date': r['updated_date'], 'name_servers': r['name_servers'],
                   'dnssec': r['dnssec'], 'resolved': bool(r['resolved']) if r['resolved'] is not None else None,
                   'block_reason': r['block_reason'], 'dns_records': r['dns_records'].split(',') if r['dns_records'] else [],
                   'error': r['error']} for r in results]
    
    if not results:
        return jsonify({'error': '没有可导出的数据'}), 400
    
    if selected:
        selected_domains = selected.split(',')
        results = [r for r in results if r['domain'] in selected_domains]
    
    if not results:
        return jsonify({'error': '没有可导出的数据'}), 400
    
    try:
        filepath, filename = create_export_file(results, format_type, filter_type)
        mimetype = 'text/csv' if format_type == 'csv' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return send_file(filepath, mimetype=mimetype, as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"导出失败: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cancel/<task_id>', methods=['POST'])
def api_cancel(task_id):
    task_pause_flags[task_id] = True
    with task_lock:
        if task_id in task_storage:
            task_storage[task_id]['status'] = 'cancelled'
            task_storage[task_id]['logs'].append({
                'time': datetime.now().strftime('%H:%M:%S'),
                'level': 'warn',
                'message': '任务已取消'
            })
    return jsonify({'message': '任务已取消'})

# ============== 历史记录API ==============

@app.route('/api/history')
def api_history():
    """获取历史记录列表"""
    limit = request.args.get('limit', 50, type=int)
    history = get_history(limit)
    return jsonify({'history': history})

@app.route('/api/history/<task_id>')
def api_history_detail(task_id):
    """获取历史详情"""
    history, results = get_history_detail(task_id)
    if not history:
        return jsonify({'error': '记录不存在'}), 404
    return jsonify({'history': history, 'results': results})

@app.route('/api/history/<task_id>', methods=['DELETE'])
def api_delete_history(task_id):
    """删除历史记录"""
    delete_history(task_id)
    return jsonify({'message': '已删除'})

@app.route('/api/history/clear', methods=['POST'])
def api_clear_history():
    """清理旧历史"""
    data = request.get_json() or {}
    days = data.get('days', 30)
    deleted = clear_old_history(days)
    return jsonify({'message': f'已清理 {deleted} 条记录'})

if __name__ == '__main__':
    print("=" * 50)
    print("域名批量查询系统 v2.3")
    print("=" * 50)
    print(f"访问地址: http://localhost:5000")
    print(f"数据目录: {DATA_DIR}")
    print(f"数据库: {DB_PATH}")
    print("=" * 50)
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
