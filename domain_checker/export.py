"""查询结果导出：CSV / XLSX。"""

import logging
import os
import tempfile
from datetime import datetime

from openpyxl import Workbook

logger = logging.getLogger(__name__)

# 导出的临时文件目录
EXPORT_DIR = tempfile.gettempdir()

EXPORT_HEADERS = ['域名', '状态', '注册商', '注册日期', '过期日期', '更新时间',
                  'DNS服务器', 'DNSSEC', '解析状态', 'DNS记录', '解析异常原因', '错误备注']

STATUS_MAP = {'success': '查询成功', 'failed': '查询失败', 'invalid': '格式无效'}
RESOLVED_MAP = {True: '正常解析', False: '未解析', None: '未知'}


def create_export_file(results: list, format: str = 'csv', filter_type: str = 'all') -> tuple:
    """按过滤条件生成导出文件，返回 (文件路径, 下载文件名)。"""
    filtered = results
    if filter_type == 'success':
        filtered = [r for r in results if r['status'] == 'success']
    elif filter_type == 'normal':
        filtered = [r for r in results
                    if r['status'] == 'success' and r.get('resolved') is not False]
    elif filter_type == 'failed':
        filtered = [r for r in results if r['status'] != 'success']
    elif filter_type == 'blocked':
        filtered = [r for r in results if r.get('resolved') is False]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if format == 'csv':
        return create_csv(filtered, timestamp), f'domain_report_{timestamp}.csv'
    return create_xlsx(filtered, timestamp), f'domain_report_{timestamp}.xlsx'


def create_csv(results: list, timestamp: str) -> str:
    """生成 CSV 报表（带 BOM，Excel 可直接打开），返回文件路径。"""
    import csv
    filepath = os.path.join(EXPORT_DIR, f'domain_report_{timestamp}.csv')

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(EXPORT_HEADERS)

        for r in results:
            dns_records = r.get('dns_records', [])
            if isinstance(dns_records, list):
                dns_records = ', '.join(dns_records)

            writer.writerow([
                r.get('domain', ''), STATUS_MAP.get(r.get('status', ''), ''),
                r.get('registrar', ''), r.get('registration_date', ''),
                r.get('expiration_date', ''), r.get('updated_date', ''),
                r.get('name_servers', ''), r.get('dnssec', ''),
                RESOLVED_MAP.get(r.get('resolved'), ''), dns_records,
                r.get('block_reason', ''), r.get('error', '')
            ])

    return filepath


def create_xlsx(results: list, timestamp: str) -> str:
    """生成 XLSX 报表，返回文件路径。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "域名查询结果"

    from openpyxl.styles import Alignment, Font, PatternFill
    for col, h in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, r in enumerate(results, 2):
        dns_records = r.get('dns_records', [])
        if isinstance(dns_records, list):
            dns_records = ', '.join(dns_records)

        for col, val in enumerate([
            r.get('domain', ''), STATUS_MAP.get(r.get('status', ''), ''),
            r.get('registrar', ''), r.get('registration_date', ''),
            r.get('expiration_date', ''), r.get('updated_date', ''),
            r.get('name_servers', ''), r.get('dnssec', ''),
            RESOLVED_MAP.get(r.get('resolved'), ''), dns_records,
            r.get('block_reason', ''), r.get('error', '')
        ], 1):
            ws.cell(row=row_idx, column=col, value=val)

    for i, w in enumerate([30, 12, 20, 12, 12, 12, 35, 15, 12, 25, 25, 35], 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

    filepath = os.path.join(EXPORT_DIR, f'domain_report_{timestamp}.xlsx')
    wb.save(filepath)
    return filepath
